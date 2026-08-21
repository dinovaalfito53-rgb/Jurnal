import openpyxl
import pandas as pd

# 1. Load Workbook
file_path = 'ReportHistory-556630.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 2. Parsing Data Transaksi
trades = []
for row in range(8, sheet.max_row + 1):
    open_time = sheet.cell(row=row, column=1).value
    if open_time in ['Orders', 'Deals', 'Working Orders', 'Summary']:
        break
        
    pos = sheet.cell(row=row, column=2).value
    symbol = sheet.cell(row=row, column=3).value
    trade_type = sheet.cell(row=row, column=4).value
    volume = sheet.cell(row=row, column=5).value
    open_price = sheet.cell(row=row, column=6).value
    sl = sheet.cell(row=row, column=7).value
    tp = sheet.cell(row=row, column=8).value
    close_time = sheet.cell(row=row, column=9).value
    close_price = sheet.cell(row=row, column=10).value
    commission = sheet.cell(row=row, column=11).value
    swap = sheet.cell(row=row, column=12).value
    profit = sheet.cell(row=row, column=13).value
    
    if open_time and close_time:
        try:
            o_dt = pd.to_datetime(str(open_time), format='%Y.%m.%d %H:%M:%S')
            c_dt = pd.to_datetime(str(close_time), format='%Y.%m.%d %H:%M:%S')
            dur_sec = (c_dt - o_dt).total_seconds()
            
            trades.append({
                'Position': pos,
                'Symbol': symbol,
                'Type': trade_type,
                'Volume': float(volume) if volume is not None else 0.0,
                'Open Time': str(open_time),
                'Open Price': float(open_price) if open_price is not None else 0.0,
                'Close Time': str(close_time),
                'Close Price': float(close_price) if close_price is not None else 0.0,
                'Commission': float(commission) if commission is not None else 0.0,
                'Swap': float(swap) if swap is not None else 0.0,
                'Profit': float(profit) if profit is not None else 0.0,
                'Duration_Sec': int(dur_sec),
                'Under_2Min': dur_sec < 120
            })
        except Exception:
            continue

df = pd.DataFrame(trades)

# 3. Urutkan: Tanggal Terbaru di Atas, Terlama di Bawah
df_sorted = df.sort_values(by='Open Time', ascending=False).reset_index(drop=True)

# 4. Hitung Metrik Keuangan
gross_profit = df_sorted[df_sorted['Profit'] > 0]['Profit'].sum()
gross_loss = abs(df_sorted[df_sorted['Profit'] < 0]['Profit'].sum())
total_commission = abs(df_sorted['Commission'].sum())
net_profit = gross_profit - gross_loss - total_commission

# 5. Generate Baris Tabel HTML dengan Highlight Kuning Terang (< 2 Min)
table_rows = ""
for idx, row in df_sorted.iterrows():
    highlight_class = 'class="fast-trade"' if row['Under_2Min'] else ''
    profit_color = '#2e7d32' if row['Profit'] >= 0 else '#c62828'
    
    table_rows += f"""
    <tr {highlight_class}>
        <td>{row['Position']}</td>
        <td><b>{row['Symbol']}</b></td>
        <td><span class="badge {row['Type']}">{row['Type'].upper()}</span></td>
        <td>{row['Volume']:.2f}</td>
        <td>{row['Open Time']}</td>
        <td>{row['Open Price']}</td>
        <td>{row['Close Time']}</td>
        <td>{row['Close Price']}</td>
        <td>${row['Commission']:.2f}</td>
        <td>${row['Swap']:.2f}</td>
        <td style="color: {profit_color}; font-weight: bold;">${row['Profit']:.2f}</td>
        <td>{row['Duration_Sec']}s ({round(row['Duration_Sec']/60, 2)}m)</td>
    </tr>
    """

# 6. Template HTML Penuh
html_content = f"""<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Trading Report - Roh_ana Phase 2</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
            background-color: #f4f6f9;
            color: #333;
            margin: 0;
            padding: 24px;
        }}
        .container {{
            max-width: 1280px;
            margin: 0 auto;
        }}
        h1 {{
            font-size: 24px;
            margin-bottom: 20px;
            color: #1a1a1a;
        }}
        .cards-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: #ffffff;
            border-radius: 8px;
            padding: 18px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            border: 1px solid #e1e4e8;
        }}
        .card-title {{
            font-size: 13px;
            color: #6c757d;
            text-transform: uppercase;
            font-weight: 600;
            margin-bottom: 8px;
        }}
        .card-value {{
            font-size: 22px;
            font-weight: bold;
        }}
        .text-green {{ color: #2e7d32; }}
        .text-red {{ color: #c62828; }}
        .text-blue {{ color: #1565c0; }}
        
        .table-container {{
            background: #ffffff;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            border: 1px solid #e1e4e8;
            overflow-x: auto;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            text-align: left;
        }}
        th {{
            background-color: #f8f9fa;
            padding: 12px 14px;
            border-bottom: 2px solid #dee2e6;
            color: #495057;
            font-weight: 600;
        }}
        td {{
            padding: 10px 14px;
            border-bottom: 1px solid #e9ecef;
        }}
        tr.fast-trade {{
            background-color: #ffff00 !important;
            color: #000000 !important;
        }}
        tr.fast-trade td {{
            border-bottom: 1px solid #d4d400;
        }}
        .badge {{
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: bold;
        }}
        .badge.buy {{ background-color: #e8f5e9; color: #2e7d32; }}
        .badge.sell {{ background-color: #ffebee; color: #c62828; }}
        .fast-trade .badge.buy {{ background-color: #1b5e20; color: #ffffff; }}
        .fast-trade .badge.sell {{ background-color: #b71c1c; color: #ffffff; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Trading History Report (Urutan: Terbaru &rarr; Terlama)</h1>
        
        <!-- Metrics Cards -->
        <div class="cards-grid">
            <div class="card">
                <div class="card-title">Gross Profit</div>
                <div class="card-value text-green">${gross_profit:.2f}</div>
            </div>
            <div class="card">
                <div class="card-title">Gross Loss</div>
                <div class="card-value text-red">-${gross_loss:.2f}</div>
            </div>
            <div class="card">
                <div class="card-title">Total Komisi</div>
                <div class="card-value text-red">-${total_commission:.2f}</div>
            </div>
            <div class="card">
                <div class="card-title">Net Profit</div>
                <div class="card-value {'text-green' if net_profit >= 0 else 'text-red'}">${net_profit:.2f}</div>
            </div>
        </div>

        <!-- Trades Table -->
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>Position</th>
                        <th>Symbol</th>
                        <th>Type</th>
                        <th>Volume</th>
                        <th>Open Time</th>
                        <th>Open Price</th>
                        <th>Close Time</th>
                        <th>Close Price</th>
                        <th>Commission</th>
                        <th>Swap</th>
                        <th>Profit</th>
                        <th>Duration</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
"""

# 7. Simpan File HTML
with open('Trading_Report_NetProfit.html', 'w', encoding='utf-8') as f:
    f.write(html_content)

print("File 'Trading_Report_NetProfit.html' berhasil dibuat.")